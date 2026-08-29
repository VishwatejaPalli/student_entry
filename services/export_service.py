"""
Export Service — Generates Excel and CSV exports.

Dynamically reads form field definitions to build column headers,
merging system fields with custom (EAV) fields.
"""

import io
import csv
import json
from datetime import datetime
from typing import Optional
import aiosqlite

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# System columns that are always exported
SYSTEM_COLUMNS = [
    ("Roll No", "roll_no"),
    ("Student Name", "student_name"),
    ("Date", "date"),
    ("Entry Time", "entry_time"),
    ("Exit Time", "exit_time"),
    ("Duration (min)", "duration_minutes"),
    ("Status", "status"),
]


async def get_export_data(
    db: aiosqlite.Connection,
    form_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
) -> tuple[list[str], list[list[str]]]:
    """
    Build export data with dynamic columns.

    Returns:
        (headers, rows) where headers is a list of column names
        and rows is a list of lists of string values.
    """
    # Get the form to export (active form if not specified)
    if form_id is None:
        cursor = await db.execute("SELECT id FROM forms WHERE active = 1 LIMIT 1")
        row = await cursor.fetchone()
        if row:
            form_id = row["id"]

    # Get custom field definitions
    custom_fields = []
    if form_id:
        cursor = await db.execute(
            """SELECT id, field_name, label, field_type FROM form_fields
               WHERE form_id = ? AND is_active = 1
               AND field_type NOT IN ('heading', 'paragraph', 'divider')
               ORDER BY position""",
            (form_id,),
        )
        custom_fields = [dict(r) for r in await cursor.fetchall()]

    # Build headers
    headers = [col[0] for col in SYSTEM_COLUMNS]
    for cf in custom_fields:
        headers.append(cf["label"] or cf["field_name"])

    # Fetch records
    conditions = []
    params = []

    if form_id:
        conditions.append("r.form_id = ?")
        params.append(form_id)
    if date_from:
        conditions.append("date(r.entry_time) >= ?")
        params.append(date_from)
    if date_to:
        conditions.append("date(r.entry_time) <= ?")
        params.append(date_to)

    where = ""
    if conditions:
        where = "WHERE " + " AND ".join(conditions)

    cursor = await db.execute(
        f"""SELECT r.*, COALESCE(s.name, '') as student_name
            FROM records r
            LEFT JOIN students s ON r.roll_no = s.roll_no
            {where}
            ORDER BY r.entry_time DESC""",
        params,
    )
    records = await cursor.fetchall()

    # Build rows
    rows = []
    for record in records:
        # System fields
        entry_dt = record["entry_time"] or ""
        date_str = ""
        entry_str = ""
        if entry_dt:
            try:
                dt = datetime.fromisoformat(entry_dt)
                date_str = dt.strftime("%d-%b-%Y")
                entry_str = dt.strftime("%H:%M")
            except (ValueError, TypeError):
                date_str = entry_dt
                entry_str = entry_dt

        exit_str = ""
        if record["exit_time"]:
            try:
                exit_str = datetime.fromisoformat(record["exit_time"]).strftime("%H:%M")
            except (ValueError, TypeError):
                exit_str = record["exit_time"]

        duration = str(record["duration_minutes"]) if record["duration_minutes"] is not None else ""
        status = "IN" if record["exit_time"] is None else "OUT"

        row = [
            record["roll_no"],
            record["student_name"],
            date_str,
            entry_str,
            exit_str,
            duration,
            status,
        ]

        # Custom fields
        if custom_fields:
            cursor2 = await db.execute(
                """SELECT field_id, value FROM record_values
                   WHERE record_id = ?""",
                (record["id"],),
            )
            values_map = {r["field_id"]: r["value"] for r in await cursor2.fetchall()}

            for cf in custom_fields:
                row.append(values_map.get(cf["id"], ""))

        rows.append(row)

    return headers, rows


async def export_excel(
    db: aiosqlite.Connection,
    form_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
) -> bytes:
    """Generate an Excel file and return bytes."""
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl is not installed")

    headers, rows = await get_export_data(db, form_id, date_from, date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Entry Records"

    # Header styling
    header_font = Font(name="Inter", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    data_font = Font(name="Inter", size=10)
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border

    # Auto-width columns
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row_data in rows:
            if col_idx - 1 < len(row_data):
                max_len = max(max_len, len(str(row_data[col_idx - 1])))
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else f"A{chr(64 + col_idx - 26)}"].width = min(max_len + 4, 40)

    # Freeze header row
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


async def export_csv(
    db: aiosqlite.Connection,
    form_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
) -> str:
    """Generate CSV string."""
    headers, rows = await get_export_data(db, form_id, date_from, date_to)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    return output.getvalue()


# ── Session Attendance Export ─────────────────────────────────────

async def export_session_excel(db: aiosqlite.Connection, session_id: int) -> bytes:
    """Generate a clean, styled Attendance Sheet Excel file for a specific class session."""
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl is not installed")

    cursor = await db.execute("SELECT * FROM class_sessions WHERE id = ?", (session_id,))
    session = await cursor.fetchone()
    if not session:
        raise ValueError("Session not found")

    cursor_st = await db.execute("""
        SELECT ss.*, COALESCE(s.name, ss.student_name, '') as full_name, s.department, s.section
        FROM session_students ss
        LEFT JOIN students s ON ss.roll_no = s.roll_no
        WHERE ss.session_id = ?
        ORDER BY ss.roll_no
    """, (session_id,))
    students = await cursor_st.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Class Attendance"

    # Fonts & Styles
    title_font = Font(name="Inter", size=14, bold=True, color="1A1A2E")
    meta_font = Font(name="Inter", size=10, color="555555")
    header_font = Font(name="Inter", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Status fills
    present_fill = PatternFill(start_color="E8F8F0", end_color="E8F8F0", fill_type="solid")
    late_fill = PatternFill(start_color="FFF8E7", end_color="FFF8E7", fill_type="solid")
    absent_fill = PatternFill(start_color="FEECEB", end_color="FEECEB", fill_type="solid")

    # Meta Header Block
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"LAB ATTENDANCE: {session['session_name'].upper()}"
    title_cell.font = title_font

    ws["A2"] = f"Class: {session['class_name'] or 'N/A'}"
    ws["C2"] = f"Subject: {session['subject'] or 'N/A'}"
    ws["E2"] = f"Room: {session['room'] or 'N/A'}"
    ws["A3"] = f"Faculty: {session['faculty'] or 'N/A'}"
    ws["C3"] = f"Date: {session['scheduled_start'][:10]}"
    ws["E3"] = f"Scheduled: {session['scheduled_start'][11:16]} - {session['scheduled_end'][11:16]}"

    for row in range(2, 4):
        for col in ["A", "C", "E"]:
            ws[f"{col}{row}"].font = meta_font

    # Headers at Row 5
    headers = [
        "S.No", "Roll Number", "Student Name", "Status",
        "PC Assigned", "Actual In", "Actual Out", "Duration (min)", "Type"
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Student Rows
    data_font = Font(name="Inter", size=10)
    for idx, s in enumerate(students, 1):
        row_num = 5 + idx
        status = s["status"]
        in_time = s["actual_entry"][11:16] if s["actual_entry"] else "—"
        out_time = s["actual_exit"][11:16] if s["actual_exit"] else "—"
        dur = str(s["duration_minutes"]) if s["duration_minutes"] is not None else "—"

        row_vals = [
            idx,
            s["roll_no"],
            s["full_name"] or s["roll_no"],
            status,
            s["pc_assigned"] or "—",
            in_time,
            out_time,
            dur,
            s["scheduled_status"],
        ]

        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = center_align if col_idx not in (2, 3) else left_align

            # Color status cell
            if col_idx == 4:
                if status == "PRESENT":
                    cell.fill = present_fill
                elif status == "LATE":
                    cell.fill = late_fill
                elif status == "ABSENT":
                    cell.fill = absent_fill

    # Summary Stats Block at the bottom
    total = len(students)
    present_cnt = sum(1 for s in students if s["status"] in ("PRESENT", "LATE"))
    late_cnt = sum(1 for s in students if s["status"] == "LATE")
    absent_cnt = sum(1 for s in students if s["status"] == "ABSENT")
    att_pct = f"{(present_cnt / total * 100):.1f}%" if total > 0 else "0%"

    sum_row = 6 + len(students) + 1
    ws.cell(row=sum_row, column=2, value=f"Total: {total}").font = meta_font
    ws.cell(row=sum_row, column=4, value=f"Present: {present_cnt} ({att_pct})").font = meta_font
    ws.cell(row=sum_row, column=6, value=f"Late: {late_cnt} | Absent: {absent_cnt}").font = meta_font

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 14

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

