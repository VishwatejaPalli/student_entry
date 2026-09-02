"""
Sessions Router — Class Sessions & Bulk Entry endpoints and pages.

Routes:
  GET  /sessions                   → Session Hub (Create Live, Bulk Log, Past Sessions)
  GET  /sessions/live/{session_id} → Live Session Cockpit
  GET  /sessions/settings          → Customizable Session Presets & Settings Page
  POST /api/sessions               → Create session (Live or Bulk Log)
  GET  /api/sessions/active        → List active sessions
  GET  /api/sessions/classes       → List classes, batches & enrolled students
  GET  /api/sessions/{id}          → Session details & student roster
  POST /api/sessions/{id}/scan     → Process barcode scan / check-in
  PUT  /api/sessions/{id}/students/{roll_no} → Update student status / PC
  POST /api/sessions/{id}/end      → End & finalize active session
  GET  /api/sessions/settings      → Get session customization presets
  POST /api/sessions/settings      → Save session customization presets
"""

from fastapi import APIRouter, Request, Query, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from typing import Optional, List
import io
import csv

from database import get_db
from models import (
    SessionCreate,
    SessionScanRequest,
    SessionStudentUpdate,
    SessionConfigModel,
    BatchAssignRequest,
)
from services.session_service import (
    create_session,
    record_session_scan,
    update_student_status,
    end_session,
    get_active_sessions,
    get_all_sessions,
    get_session_details,
    get_classes_list,
    get_session_config,
    save_session_config,
)

router = APIRouter()
templates = Jinja2Templates(directory="templates")


# ── Pages ─────────────────────────────────────────────────────────

@router.get("/sessions", response_class=HTMLResponse)
async def sessions_hub_page(request: Request):
    """Session Hub: Start Live Class, Bulk Direct Entry, and Past Sessions."""
    db = await get_db()
    try:
        classes = await get_classes_list(db)
        active_sessions = await get_active_sessions(db)
        past_sessions = await get_all_sessions(db, limit=20)
        session_config = await get_session_config(db)

        departments = sorted(list(set(c["department"] for c in classes if c.get("department"))))

        return templates.TemplateResponse(request, name="sessions/index.html", context={
            "classes": classes,
            "departments": departments,
            "active_sessions": active_sessions,
            "past_sessions": past_sessions,
            "session_config": session_config,
        })
    finally:
        await db.close()


@router.get("/sessions/live/{session_id}", response_class=HTMLResponse)
async def live_session_page(request: Request, session_id: int):
    """Live Session Cockpit with real-time barcode scanner and roster."""
    db = await get_db()
    try:
        session = await get_session_details(db, session_id)
        if not session:
            return HTMLResponse("Session not found", status_code=404)

        return templates.TemplateResponse(request, name="sessions/live.html", context={
            "session": session,
        })
    finally:
        await db.close()


@router.get("/sessions/settings", response_class=HTMLResponse)
async def session_settings_page(request: Request):
    """Customizable Bulk Entry & Session Presets/Settings page."""
    db = await get_db()
    try:
        session_config = await get_session_config(db)
        return templates.TemplateResponse(request, name="sessions/settings.html", context={
            "session_config": session_config,
        })
    finally:
        await db.close()


# ── API Endpoints ─────────────────────────────────────────────────

@router.get("/api/sessions/classes")
async def api_get_classes():
    """Get list of classes, batches, and student rosters."""
    db = await get_db()
    try:
        classes = await get_classes_list(db)
        return JSONResponse(content=classes)
    finally:
        await db.close()


@router.get("/api/sessions/settings")
async def api_get_session_settings():
    """Get customizable session presets and options."""
    db = await get_db()
    try:
        config = await get_session_config(db)
        return JSONResponse(content=config)
    finally:
        await db.close()


@router.post("/api/sessions/settings")
async def api_save_session_settings(data: SessionConfigModel):
    """Save customizable session presets, batch lists, and dynamic fields."""
    db = await get_db()
    try:
        saved = await save_session_config(db, data.model_dump())
        return JSONResponse({
            "success": True,
            "message": "Bulk session settings saved successfully",
            "config": saved,
        })
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=400)
    finally:
        await db.close()


@router.get("/api/sessions/active")
async def api_get_active_sessions():
    """Get list of currently active sessions."""
    db = await get_db()
    try:
        active = await get_active_sessions(db)
        return JSONResponse(content=active)
    finally:
        await db.close()


@router.get("/api/sessions/{session_id}")
async def api_get_session(session_id: int):
    """Get full details and roster for a session."""
    db = await get_db()
    try:
        session = await get_session_details(db, session_id)
        if not session:
            return JSONResponse({"error": "Session not found"}, status_code=404)
        return JSONResponse(content=session)
    finally:
        await db.close()


@router.post("/api/sessions")
async def api_create_session(data: SessionCreate):
    """Create a new class session (Live or Immediate Bulk Record)."""
    db = await get_db()
    try:
        session_id = await create_session(
            db=db,
            session_data=data.model_dump(),
            student_roll_nos=data.students,
            is_completed_bulk=data.is_completed_bulk,
            bulk_status=data.bulk_status,
        )
        return JSONResponse({
            "success": True,
            "session_id": session_id,
            "message": "Class session created successfully",
            "redirect_url": f"/sessions/live/{session_id}" if not data.is_completed_bulk else "/sessions",
        }, status_code=201)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=400)
    finally:
        await db.close()


@router.post("/api/sessions/{session_id}/scan")
async def api_session_scan(session_id: int, data: SessionScanRequest):
    """Process a barcode scan or manual roll entry for an active class session."""
    db = await get_db()
    try:
        result = await record_session_scan(
            db=db,
            session_id=session_id,
            roll_no=data.roll_no,
            manual_pc=data.pc_assigned,
        )
        return JSONResponse(content=result)
    finally:
        await db.close()


@router.put("/api/sessions/{session_id}/students/{roll_no}")
async def api_update_student_in_session(
    session_id: int,
    roll_no: str,
    data: SessionStudentUpdate
):
    """Manually update student status (Present/Absent/Late) or assigned PC."""
    db = await get_db()
    try:
        ok = await update_student_status(
            db=db,
            session_id=session_id,
            roll_no=roll_no,
            status=data.status,
            pc_assigned=data.pc_assigned,
            actual_entry=data.actual_entry,
            actual_exit=data.actual_exit,
        )
        if not ok:
            return JSONResponse({"error": "No updates specified"}, status_code=400)
        return JSONResponse({"success": True, "message": "Student updated"})
    finally:
        await db.close()


@router.post("/api/sessions/{session_id}/end")
async def api_end_session(session_id: int):
    """End active session: marks un-scanned students as ABSENT, calculates all durations."""
    db = await get_db()
    try:
        result = await end_session(db, session_id)
        return JSONResponse({
            "success": True,
            "message": "Session finalized successfully",
            "session": result,
        })
    finally:
        await db.close()


@router.post("/api/sessions/parse-file")
async def api_parse_session_file(file: UploadFile = File(...)):
    """Parse an uploaded Excel or CSV file to extract roll numbers for bulk selection."""
    try:
        content = await file.read()
        roll_numbers = []

        if file.filename and file.filename.endswith(('.xlsx', '.xls')):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content))
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        val = str(cell).strip().upper()
                        # Match alphanumeric roll number tokens
                        if len(val) >= 4 and any(c.isdigit() for c in val):
                            roll_numbers.append(val)
                            break
        else:
            text = content.decode("utf-8", errors="ignore")
            reader = csv.reader(io.StringIO(text))
            for row in reader:
                if row:
                    for cell in row:
                        val = cell.strip().upper()
                        if len(val) >= 4 and any(c.isdigit() for c in val):
                            roll_numbers.append(val)
                            break

        # Deduplicate preserving order
        seen = set()
        unique_rolls = []
        for r in roll_numbers:
            if r not in seen and r not in ("ROLL_NO", "ROLL NO", "ROLLNO", "STUDENT ROLL NO"):
                seen.add(r)
                unique_rolls.append(r)

        return JSONResponse({
            "success": True,
            "roll_numbers": unique_rolls,
            "count": len(unique_rolls),
        })
    except Exception as e:
        return JSONResponse({"error": f"Failed to parse file: {str(e)}"}, status_code=400)


@router.post("/api/sessions/assign-batches")
async def api_assign_batches(data: BatchAssignRequest):
    """Permanently divide and allocate class students into semester batches in the database."""
    from services.session_service import assign_batches_to_class
    db = await get_db()
    try:
        result = await assign_batches_to_class(
            db=db,
            class_name=data.class_name,
            split_count=data.split_count,
            prefix=data.prefix,
            ranges=data.ranges,
        )
        return JSONResponse(result)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=400)
    finally:
        await db.close()
