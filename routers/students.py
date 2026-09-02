"""
Students Router — Student management CRUD with batch support.

Routes:
  GET    /admin/students          → Student management page
  POST   /api/students            → Add single student
  PUT    /api/students/{id}       → Edit student
  POST   /api/students/import     → Import from Excel/CSV
  GET    /api/students/search     → Search students
  GET    /api/students            → List all students (paginated + filtered)
"""

import csv
import io
import os
from fastapi import APIRouter, Request, Query, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse, Response
from fastapi.templating import Jinja2Templates
from typing import Optional

from database import get_db
from models import StudentCreate, StudentUpdate

router = APIRouter()
templates = Jinja2Templates(directory="templates")


@router.get("/admin/students", response_class=HTMLResponse)
async def students_page(request: Request):
    """Student management page."""
    db = await get_db()
    try:
        cursor = await db.execute(
            "SELECT * FROM students WHERE active = 1 ORDER BY department, section, roll_no LIMIT 1000"
        )
        students = [dict(r) for r in await cursor.fetchall()]

        cursor = await db.execute("SELECT COUNT(*) as cnt FROM students WHERE active = 1")
        row = await cursor.fetchone()
        total = row["cnt"]

        # Get unique departments, sections, batches for filtering
        cursor_dept = await db.execute("SELECT DISTINCT department FROM students WHERE active = 1 AND department != '' ORDER BY department")
        departments = [r["department"] for r in await cursor_dept.fetchall()]

        cursor_batch = await db.execute("SELECT DISTINCT batch FROM students WHERE active = 1 AND batch != '' ORDER BY batch")
        batches = [r["batch"] for r in await cursor_batch.fetchall()]

        return templates.TemplateResponse(request, name="admin/students.html", context={
            "students": students,
            "total": total,
            "departments": departments,
            "batches": batches,
        })
    finally:
        await db.close()


@router.get("/api/students/sample-template")
async def api_download_sample_template():
    """Download sample CSV template for student import."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["roll_no", "name", "department", "section", "batch", "year"])
    writer.writerow(["24885A0401", "Aarav Sharma", "ECE", "A", "A1", "3"])
    writer.writerow(["24885A0402", "Aditi Patel", "ECE", "A", "A1", "3"])
    writer.writerow(["24885A0406", "Deepak Verma", "ECE", "A", "A2", "3"])
    writer.writerow(["24885A0411", "Kavya Iyer", "ECE", "B", "B1", "3"])
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=students_import_template.csv"}
    )


@router.get("/api/students")
async def api_list_students(
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    search: Optional[str] = None,
    department: Optional[str] = None,
    batch: Optional[str] = None,
):
    """List students with optional search, department, and batch filtering."""
    db = await get_db()
    try:
        conditions = ["active = 1"]
        params = []

        if search:
            conditions.append("(roll_no LIKE ? OR name LIKE ?)")
            params.extend([f"%{search}%", f"%{search}%"])

        if department:
            conditions.append("department = ?")
            params.append(department)

        if batch:
            conditions.append("batch = ?")
            params.append(batch)

        where = "WHERE " + " AND ".join(conditions)

        cursor = await db.execute(f"SELECT COUNT(*) as cnt FROM students {where}", params)
        row = await cursor.fetchone()
        total = row["cnt"]

        offset = (page - 1) * per_page
        cursor = await db.execute(
            f"""SELECT * FROM students {where}
                ORDER BY department, section, roll_no
                LIMIT ? OFFSET ?""",
            params + [per_page, offset],
        )
        students = [dict(r) for r in await cursor.fetchall()]

        return JSONResponse({
            "students": students,
            "total": total,
            "page": page,
            "pages": (total + per_page - 1) // per_page,
        })
    finally:
        await db.close()


@router.get("/api/students/search")
async def api_search_students(q: str = Query("", min_length=1)):
    """Search students by roll number or name."""
    db = await get_db()
    try:
        cursor = await db.execute(
            """SELECT id, roll_no, name, department, section, batch, year
               FROM students
               WHERE active = 1 AND (roll_no LIKE ? OR name LIKE ?)
               ORDER BY roll_no
               LIMIT 20""",
            (f"%{q.upper()}%", f"%{q}%"),
        )
        results = [dict(r) for r in await cursor.fetchall()]
        return JSONResponse(content=results)
    finally:
        await db.close()


@router.post("/api/students")
async def api_create_student(data: StudentCreate):
    """Add a single student."""
    db = await get_db()
    try:
        roll_no = data.roll_no.strip().upper()

        # Check for duplicate
        cursor = await db.execute(
            "SELECT id FROM students WHERE roll_no = ?", (roll_no,)
        )
        existing = await cursor.fetchone()
        if existing:
            return JSONResponse(
                {"error": f"Student {roll_no} already exists"},
                status_code=409,
            )

        await db.execute(
            """INSERT INTO students (roll_no, name, department, section, batch, year)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (roll_no, data.name.strip(), data.department.strip(), data.section.strip(), data.batch.strip().upper(), data.year.strip()),
        )
        await db.commit()
        return JSONResponse({"success": True, "roll_no": roll_no}, status_code=201)
    finally:
        await db.close()


@router.put("/api/students/{student_id}")
async def api_update_student(student_id: int, data: StudentUpdate):
    """Update student details."""
    db = await get_db()
    try:
        updates = []
        params = []

        if data.name is not None:
            updates.append("name = ?")
            params.append(data.name.strip())
        if data.department is not None:
            updates.append("department = ?")
            params.append(data.department.strip())
        if data.section is not None:
            updates.append("section = ?")
            params.append(data.section.strip())
        if data.batch is not None:
            updates.append("batch = ?")
            params.append(data.batch.strip().upper())
        if data.year is not None:
            updates.append("year = ?")
            params.append(data.year.strip())
        if data.active is not None:
            updates.append("active = ?")
            params.append(1 if data.active else 0)

        if not updates:
            return JSONResponse({"error": "No fields to update"}, status_code=400)

        params.append(student_id)
        await db.execute(
            f"UPDATE students SET {', '.join(updates)} WHERE id = ?",
            params,
        )
        await db.commit()
        return JSONResponse({"success": True})
    finally:
        await db.close()


@router.delete("/api/students/{student_id}")
async def api_delete_student(student_id: int):
    """Delete a student from the directory with cascade cleanup."""
    db = await get_db()
    try:
        cursor = await db.execute("SELECT id, roll_no, name FROM students WHERE id = ?", (student_id,))
        student = await cursor.fetchone()
        if not student:
            return JSONResponse({"error": "Student not found"}, status_code=404)

        roll_no = student["roll_no"]

        # 1. Clean up session students mappings
        await db.execute("DELETE FROM session_students WHERE roll_no = ?", (roll_no,))

        # 2. Clean up records and associated record values
        cursor_rec = await db.execute("SELECT id FROM records WHERE roll_no = ?", (roll_no,))
        rec_rows = await cursor_rec.fetchall()
        for rec in rec_rows:
            await db.execute("DELETE FROM record_values WHERE record_id = ?", (rec["id"],))
        await db.execute("DELETE FROM records WHERE roll_no = ?", (roll_no,))

        # 3. Delete student
        await db.execute("DELETE FROM students WHERE id = ?", (student_id,))
        await db.commit()

        return JSONResponse({
            "success": True,
            "message": f"Student {roll_no} ({student['name'] or 'No Name'}) removed successfully."
        })
    finally:
        await db.close()


@router.post("/api/students/import")
async def api_import_students(file: UploadFile = File(...)):
    """
    Import students from CSV or Excel.
    Expected columns: roll_no, name, department, section, batch, year
    """
    db = await get_db()
    try:
        content = await file.read()

        if file.filename and file.filename.endswith(('.xlsx', '.xls')):
            # Excel import
            try:
                from openpyxl import load_workbook
                wb = load_workbook(io.BytesIO(content))
                ws = wb.active
                rows_iter = ws.iter_rows(values_only=True)
                header = [str(h).strip().lower() if h else "" for h in next(rows_iter)]
                data_rows = []
                for row in rows_iter:
                    data_rows.append({header[i]: str(row[i]).strip() if row[i] else "" for i in range(len(header))})
            except Exception as e:
                return JSONResponse({"error": f"Failed to parse Excel: {str(e)}"}, status_code=400)
        else:
            # CSV import
            try:
                text = content.decode("utf-8")
                reader = csv.DictReader(io.StringIO(text))
                data_rows = []
                for row in reader:
                    cleaned = {k.strip().lower(): v.strip() for k, v in row.items()}
                    data_rows.append(cleaned)
            except Exception as e:
                return JSONResponse({"error": f"Failed to parse CSV: {str(e)}"}, status_code=400)

        imported = 0
        skipped = 0
        errors_list = []

        for row in data_rows:
            roll_no = row.get("roll_no", row.get("roll no", row.get("rollno", ""))).strip().upper()
            if not roll_no:
                skipped += 1
                continue

            name = row.get("name", row.get("student name", ""))
            department = row.get("department", row.get("dept", ""))
            section = row.get("section", row.get("sec", ""))
            batch = row.get("batch", row.get("sub_batch", "")).strip().upper()
            year = row.get("year", row.get("yr", ""))

            try:
                cursor = await db.execute(
                    "SELECT id FROM students WHERE roll_no = ?", (roll_no,)
                )
                existing = await cursor.fetchone()

                if existing:
                    # Update existing
                    await db.execute(
                        """UPDATE students SET name = ?, department = ?, section = ?, batch = ?, year = ?
                           WHERE roll_no = ?""",
                        (name, department, section, batch, year, roll_no),
                    )
                else:
                    await db.execute(
                        """INSERT INTO students (roll_no, name, department, section, batch, year)
                           VALUES (?, ?, ?, ?, ?, ?)""",
                        (roll_no, name, department, section, batch, year),
                    )
                imported += 1
            except Exception as e:
                errors_list.append(f"{roll_no}: {str(e)}")
                skipped += 1

        await db.commit()

        return JSONResponse({
            "success": True,
            "imported": imported,
            "skipped": skipped,
            "errors": errors_list[:10],
        })
    finally:
        await db.close()


@router.get("/api/admin/data-summary")
async def api_data_summary():
    """Returns total counts of students, records, and sessions for confirmation before clear."""
    db = await get_db()
    try:
        c1 = await db.execute("SELECT COUNT(*) as cnt FROM students")
        students_count = (await c1.fetchone())["cnt"]

        c2 = await db.execute("SELECT COUNT(*) as cnt FROM records")
        records_count = (await c2.fetchone())["cnt"]

        c3 = await db.execute("SELECT COUNT(*) as cnt FROM class_sessions")
        sessions_count = (await c3.fetchone())["cnt"]

        return JSONResponse({
            "students": students_count,
            "records": records_count,
            "sessions": sessions_count,
        })
    finally:
        await db.close()


@router.post("/api/admin/clear-data")
async def api_clear_data(request: Request):
    """
    Password-protected clear / reset data in one go for year or semester transition.
    Body JSON:
      - password: str
      - target: 'students' | 'logs' | 'all'
    """
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"error": "Invalid request body"}, status_code=400)

    password = str(body.get("password", "")).strip()
    target = str(body.get("target", "all")).strip()

    # Dynamically read ADMIN_PASSWORD from .env so any changes take effect immediately
    from dotenv import dotenv_values
    from pathlib import Path
    env_cfg = dotenv_values(Path(__file__).resolve().parent.parent / ".env")
    admin_password = env_cfg.get("ADMIN_PASSWORD") or os.environ.get("ADMIN_PASSWORD", "admin")

    if password != admin_password:
        return JSONResponse({"error": "Incorrect admin password. Action aborted."}, status_code=401)

    db = await get_db()
    try:
        deleted_summary = {}

        if target in ("students", "all"):
            # Cascade delete records and session mappings first
            await db.execute("DELETE FROM record_values")
            await db.execute("DELETE FROM session_students")
            await db.execute("DELETE FROM records")
            
            c_stud = await db.execute("SELECT COUNT(*) as cnt FROM students")
            stud_cnt = (await c_stud.fetchone())["cnt"]
            await db.execute("DELETE FROM students")
            deleted_summary["students"] = stud_cnt
            deleted_summary["records"] = "All"

        if target == "logs":
            # Delete only entry logs & class sessions, keep students intact
            await db.execute("DELETE FROM record_values")
            await db.execute("DELETE FROM session_students")
            await db.execute("DELETE FROM records")
            await db.execute("DELETE FROM class_sessions")
            deleted_summary["records"] = "All"
            deleted_summary["sessions"] = "All"

        if target == "all":
            await db.execute("DELETE FROM class_sessions")
            deleted_summary["sessions"] = "All"

        await db.commit()

        target_desc = {
            "students": "Student Master Directory & associated records",
            "logs": "Activity Logs & Class Sessions",
            "all": "All Student Profiles, Records, and Class Sessions",
        }.get(target, "Selected data")

        return JSONResponse({
            "success": True,
            "message": f"Successfully cleared {target_desc} in one go.",
            "summary": deleted_summary,
        })
    finally:
        await db.close()
